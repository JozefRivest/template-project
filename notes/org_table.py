#!/usr/bin/env python3
"""
org_table.py - Render org/YAML reading notes as a table

Org file structure:
  * Section Name          <- becomes a full-width separator row in the table
  ** Author Year          <- becomes an entry (Author field)
  :PROPERTIES:
  :Date: 2015
  :Journal: JJSS
  :END:

YAML file structure:
  - category: Section Name
    entries:
      - author: Author Year
        date: 2015
        journal: JJSS
        ...
    subcategories:
      - subcategory: Subsection Name
        entries:
          - author: Author Year
            date: 2015
            ...

Usage:
    python3 org_table.py readings.yaml
    python3 org_table.py readings.yaml
    python3 org_table.py readings.yaml --html
    python3 org_table.py readings.yaml --graph
    python3 org_table.py readings.yaml --live
    python3 org_table.py readings.yaml --columns Author Date Journal Claim
    python3 org_table.py readings.yaml --sort Date
    python3 org_table.py readings.yaml --filter Journal=APSR
"""

import argparse
import os
import re
import tempfile
import threading
import time
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

# ── Parsing ────────────────────────────────────────────────────────────────────


def parse_yaml(filepath: str) -> list[dict]:
    try:
        import yaml
    except ImportError:
        raise SystemExit("PyYAML is required for YAML files: pip install pyyaml")

    with open(filepath, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    if not isinstance(data, list):
        raise SystemExit(
            "YAML file must be a list of category objects at the top level."
        )

    def parse_entries(entries, category, subcategory=""):
        rows = []
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            row = {
                "_type": "entry",
                "_section": str(category),
                "_subsection": str(subcategory),
            }
            for k, v in entry.items():
                row[k.capitalize() if k == "author" else k] = (
                    str(v) if v is not None else ""
                )
            if "Author" not in row:
                row["Author"] = ""
            rows.append(row)
        return rows

    items = []
    for block in data:
        if not isinstance(block, dict):
            continue
        category = block.get("category", "")
        if category:
            items.append({"_type": "section", "title": str(category)})
        for key, value in block.items():
            if key == "entries":
                items.extend(parse_entries(value or [], category))
            elif key == "subcategories":
                for subblock in value or []:
                    if not isinstance(subblock, dict):
                        continue
                    subcategory = subblock.get("subcategory", "")
                    if subcategory:
                        items.append({"_type": "subsection", "title": str(subcategory)})
                    items.extend(
                        parse_entries(
                            subblock.get("entries", []), category, subcategory
                        )
                    )
    return items


def parse_org(filepath: str) -> list[dict]:
    items = []
    current_entry = None
    in_properties = False
    current_section = None

    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        line = line.rstrip("\n")

        if re.match(r"^\* [^*]", line):
            if current_entry is not None:
                items.append(current_entry)
                current_entry = None
            title = re.sub(r"^\* ", "", line).strip()
            current_section = title
            items.append({"_type": "section", "title": title})
            in_properties = False
            continue

        if re.match(r"^\*{2,} ", line):
            if current_entry is not None:
                items.append(current_entry)
            heading = re.sub(r"^\*+ ", "", line).strip()
            current_entry = {
                "_type": "entry",
                "_section": current_section,
                "Author": heading,
            }
            in_properties = False
            continue

        if current_entry is None:
            continue

        if line.strip() == ":PROPERTIES:":
            in_properties = True
            continue

        if line.strip() == ":END:":
            in_properties = False
            continue

        if in_properties:
            match = re.match(r"^:([^:]+):\s*(.*)", line.strip())
            if match:
                key, value = match.group(1).strip(), match.group(2).strip()
                current_entry[key] = value

    if current_entry is not None:
        items.append(current_entry)

    return items


# ── Helpers ────────────────────────────────────────────────────────────────────


def entries_only(items):
    return [i for i in items if i.get("_type") == "entry"]


def apply_filters(items, sort_by, filter_by):
    entries = entries_only(items)

    if not entries:
        print("No entries found.")
        return None, None

    if filter_by:
        key, _, value = filter_by.partition("=")
        entries = [e for e in entries if e.get(key, "").lower() == value.lower()]
        if not entries:
            print(f"No entries match {filter_by}")
            return None, None

    if filter_by or sort_by:
        if sort_by:
            entries = sorted(entries, key=lambda e: e.get(sort_by, ""))
        return entries, False

    return items, True


def resolve_columns(entries, columns):
    if columns:
        return columns
    seen = {}
    for e in entries:
        for k in e:
            if not k.startswith("_"):
                seen[k] = True
    return list(seen.keys())


def esc(text: str) -> str:
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def strip_inline(text: str) -> str:
    """Strip **bold** and *italic*/_italic_ markers from plain-text output."""
    s = str(text)
    s = re.sub(r"\*\*(.+?)\*\*", r"\1", s, flags=re.DOTALL)
    s = re.sub(r"\*(.+?)\*", r"\1", s, flags=re.DOTALL)
    s = re.sub(r"_(.+?)_", r"\1", s, flags=re.DOTALL)
    return s


def render_inline(text: str) -> str:
    """Convert **bold** and *italic*/_italic_ markers to HTML, escaping everything else."""
    parts = re.split(r"(\*\*[^*\n]+?\*\*|\*[^*\n]+?\*|_[^_\n]+?_)", str(text))
    result = []
    for part in parts:
        if part.startswith("**") and part.endswith("**") and len(part) > 4:
            result.append(f"<strong>{esc(part[2:-2])}</strong>")
        elif (part.startswith("*") and part.endswith("*") and len(part) > 2) or (
            part.startswith("_") and part.endswith("_") and len(part) > 2
        ):
            result.append(f"<em>{esc(part[1:-1])}</em>")
        else:
            result.append(esc(part))
    return "".join(result)


# ── HTML builder ───────────────────────────────────────────────────────────────


def build_html(items, filepath, columns=None, sort_by=None, filter_by=None, live=False):
    result, sectioned = apply_filters(items, sort_by, filter_by)
    if result is None:
        result = []
        sectioned = False

    all_entries = entries_only(result) if sectioned else result
    cols = resolve_columns(all_entries, columns)
    title = Path(filepath).stem
    num_cols = len(cols)

    header_cells = "".join(f"<th>{esc(col)}</th>" for col in cols)

    body_rows = ""
    if sectioned:
        for item in result:
            if item["_type"] == "section":
                body_rows += (
                    f'<tr class="section-row">'
                    f'<td colspan="{num_cols}">{esc(item["title"])}</td>'
                    f"</tr>\n"
                )
            elif item["_type"] == "subsection":
                body_rows += (
                    f'<tr class="subsection-row">'
                    f'<td colspan="{num_cols}">{esc(item["title"])}</td>'
                    f"</tr>\n"
                )
            else:
                cells = "".join(
                    f"<td>{render_inline(item.get(col, ''))}</td>" for col in cols
                )
                body_rows += f"<tr>{cells}</tr>\n"
    else:
        for e in result:
            cells = "".join(f"<td>{render_inline(e.get(col, ''))}</td>" for col in cols)
            body_rows += f"<tr>{cells}</tr>\n"

    entry_count = len(all_entries)

    live_script = (
        """
    const evtSource = new EventSource("/sse");
    evtSource.onmessage = function(e) {
      if (e.data === "reload") {
        fetch("/table")
          .then(r => r.text())
          .then(html => {
            const parser = new DOMParser();
            const doc = parser.parseFromString(html, "text/html");
            document.querySelector("tbody").innerHTML =
              doc.querySelector("tbody").innerHTML;
            document.getElementById("count").textContent =
              doc.getElementById("count").textContent;
          });
      }
    };
    """
        if live
        else ""
    )

    live_badge = '<span class="live-badge">● LIVE</span>' if live else ""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{esc(title)}</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
      font-size: 18px;
      background: #f5f5f5;
      color: #222;
      padding: 2rem;
    }}
    .header {{
      display: flex;
      align-items: center;
      gap: 0.75rem;
      margin-bottom: 1rem;
    }}
    h1 {{ font-size: 1.4rem; font-weight: 600; color: #333; }}
    .live-badge {{
      font-size: 0.75rem;
      font-weight: 600;
      color: #2d6a4f;
      background: #d8f3dc;
      padding: 2px 8px;
      border-radius: 999px;
      animation: pulse 2s infinite;
    }}
    @keyframes pulse {{
      0%, 100% {{ opacity: 1; }}
      50% {{ opacity: 0.4; }}
    }}
    .controls {{ display: flex; gap: 0.75rem; margin-bottom: 1rem; }}
    input[type="text"] {{
      padding: 7px 12px;
      border: 1px solid #ccc;
      border-radius: 6px;
      font-size: 14px;
      width: 300px;
      outline: none;
    }}
    input[type="text"]:focus {{ border-color: #2d6a4f; }}
    .table-wrapper {{
      overflow-x: auto;
      border-radius: 8px;
      box-shadow: 0 1px 4px rgba(0,0,0,0.12);
    }}
    table {{ border-collapse: collapse; width: 100%; background: #fff; }}
    thead {{ background: #2d6a4f; color: #fff; }}
    th {{
      padding: 10px 14px;
      text-align: left;
      font-weight: 600;
      white-space: nowrap;
      cursor: pointer;
      user-select: none;
    }}
    th:hover {{ background: #245a42; }}
    th.sorted-asc::after  {{ content: " ▲"; font-size: 0.75em; }}
    th.sorted-desc::after {{ content: " ▼"; font-size: 0.75em; }}
    td {{
      padding: 9px 14px;
      border-bottom: 1px solid #eee;
      vertical-align: top;
      max-width: 300px;
      word-wrap: break-word;
    }}
    tr.section-row td {{
      background: #b7e4c7;
      color: #1b4332;
      font-weight: 700;
      font-size: 0.9rem;
      letter-spacing: 0.04em;
      text-transform: uppercase;
      padding: 7px 14px;
      border-bottom: 2px solid #74c69d;
    }}
    tr.subsection-row td {{
      background: #d8f3dc;
      color: #2d6a4f;
      font-weight: 600;
      font-size: 0.85rem;
      letter-spacing: 0.02em;
      padding: 5px 20px;
      border-bottom: 1px solid #95d5b2;
    }}
    tr:last-child td {{ border-bottom: none; }}
    tr:not(.section-row):nth-child(even) {{ background: #f9f9f9; }}
    tr:not(.section-row):hover {{ background: #eaf4ee; }}
    .count {{ margin-top: 0.75rem; font-size: 0.85rem; color: #888; }}
  </style>
</head>
<body>
  <div class="header">
    <h1>{esc(title)}</h1>
    {live_badge}
  </div>
  <div class="controls">
    <input type="text" id="search" placeholder="Search entries…" oninput="filterTable()" />
  </div>
  <div class="table-wrapper">
    <table id="main-table">
      <thead><tr>{header_cells}</tr></thead>
      <tbody>{body_rows}</tbody>
    </table>
  </div>
  <p class="count" id="count">{entry_count} entries</p>

  <script>
    function filterTable() {{
      const query = document.getElementById("search").value.toLowerCase();
      const rows = document.querySelectorAll("#main-table tbody tr");
      let visible = 0;
      let lastSection = null;
      rows.forEach(row => {{
        if (row.classList.contains("section-row")) {{
          lastSection = row;
          row.style.display = "none";
          return;
        }}
        const match = row.textContent.toLowerCase().includes(query);
        row.style.display = match ? "" : "none";
        if (match) {{
          visible++;
          if (lastSection) {{ lastSection.style.display = ""; lastSection = null; }}
        }}
      }});
      document.getElementById("count").textContent = visible + " entries";
    }}

    let sortCol = -1, sortAsc = true;
    document.querySelectorAll("#main-table thead th").forEach((th, idx) => {{
      th.addEventListener("click", () => {{
        const tbody = document.querySelector("#main-table tbody");
        const rows = Array.from(tbody.querySelectorAll("tr:not(.section-row)"));
        sortAsc = sortCol === idx ? !sortAsc : true;
        sortCol = idx;
        rows.sort((a, b) => {{
          const aText = a.cells[idx]?.textContent.trim() ?? "";
          const bText = b.cells[idx]?.textContent.trim() ?? "";
          return sortAsc ? aText.localeCompare(bText) : bText.localeCompare(aText);
        }});
        rows.forEach(r => tbody.appendChild(r));
        document.querySelectorAll("#main-table thead th").forEach(h =>
          h.classList.remove("sorted-asc", "sorted-desc"));
        th.classList.add(sortAsc ? "sorted-asc" : "sorted-desc");
      }});
    }});

    {live_script}
  </script>
</body>
</html>"""


# ── Citation graph ─────────────────────────────────────────────────────────────


def get_doi(entry: dict) -> str:
    """Return a normalized DOI for an entry, or '' if none present."""
    raw = ""
    for key in ("doi", "DOI", "Doi"):
        if entry.get(key):
            raw = entry[key]
            break
    raw = str(raw).strip()
    if not raw:
        return ""
    raw = re.sub(r"^https?://(dx\.)?doi\.org/", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"^doi:\s*", "", raw, flags=re.IGNORECASE)
    return raw.strip().lower()


def graph_cache_path(filepath: str) -> Path:
    return Path(filepath).resolve().parent / ".graph_cache.json"


def load_graph_cache(path: Path) -> dict:
    import json

    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def save_graph_cache(path: Path, cache: dict) -> None:
    import json

    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2)
    except OSError as e:
        print(f"Warning: could not write graph cache: {e}")


S2_REFERENCES_URL = (
    "https://api.semanticscholar.org/graph/v1/paper/DOI:{doi}/references"
    "?fields=externalIds,title,year,authors&limit=1000"
)


def _http_json(url: str, headers: dict):
    """GET `url` and return parsed JSON, or None on 404/400/connection error."""
    import json
    from urllib.error import HTTPError, URLError
    from urllib.request import Request, urlopen

    for attempt in range(5):
        try:
            with urlopen(Request(url, headers=headers), timeout=30) as resp:
                return json.load(resp)
        except HTTPError as e:
            if e.code == 429 and attempt < 4:
                retry_after = e.headers.get("Retry-After")
                wait = (
                    int(retry_after)
                    if retry_after and retry_after.isdigit()
                    else 5 * (attempt + 1)
                )
                time.sleep(wait)
                continue
            if e.code in (404, 400):
                return None
            raise
        except URLError:
            raise
    return None


def _fetch_references_s2(doi: str) -> list[dict]:
    from urllib.parse import quote

    headers = {"User-Agent": "reading-notes-table/graph"}
    api_key = os.environ.get("S2_API_KEY")
    if api_key:
        headers["x-api-key"] = api_key
    data = _http_json(S2_REFERENCES_URL.format(doi=quote(doi, safe="")), headers)
    if not data:
        return []
    refs = []
    for item in data.get("data") or []:
        cited = item.get("citedPaper") or {}
        ext = cited.get("externalIds") or {}
        authors = cited.get("authors") or []
        refs.append(
            {
                "doi": str(ext.get("DOI") or "").lower(),
                "title": cited.get("title") or "",
                "year": cited.get("year"),
                "author": authors[0].get("name", "") if authors else "",
            }
        )
    return refs


def _fetch_references_crossref(doi: str) -> list[dict]:
    from urllib.parse import quote

    headers = {"User-Agent": "reading-notes-table/graph (citation graph)"}
    data = _http_json(f"https://api.crossref.org/works/{quote(doi, safe='')}", headers)
    if not data:
        return []
    refs = []
    for r in (data.get("message") or {}).get("reference") or []:
        rdoi = str(r.get("DOI") or "").lower()
        year = r.get("year")
        try:
            year = int(year) if year else None
        except (ValueError, TypeError):
            year = None
        refs.append(
            {
                "doi": rdoi,
                "title": r.get("article-title") or r.get("volume-title") or "",
                "year": year,
                "author": r.get("author") or "",
            }
        )
    return refs


def _fetch_references_opencitations(doi: str) -> list[dict]:
    from urllib.parse import quote

    headers = {"User-Agent": "reading-notes-table/graph"}
    data = _http_json(
        f"https://opencitations.net/index/api/v2/references/doi:{quote(doi, safe='')}",
        headers,
    )
    if not data:
        return []
    refs = []
    for r in data:
        cited = str(r.get("cited") or "")
        rdoi = ""
        for token in cited.split():
            if token.startswith("doi:"):
                rdoi = token[4:].lower()
                break
        if rdoi:
            refs.append({"doi": rdoi, "title": "", "year": None, "author": ""})
    return refs


def fetch_references(doi: str) -> tuple[list[dict], str]:
    """Fetch papers cited by `doi`, trying several sources. Returns (refs, source)."""
    for name, fn in (
        ("semanticscholar", _fetch_references_s2),
        ("crossref", _fetch_references_crossref),
        ("opencitations", _fetch_references_opencitations),
    ):
        refs = fn(doi)
        if refs:
            return refs, name
    return [], "none"


def fetch_all_references(items, cache, offline=False):
    """Populate `cache` with references for every entry DOI. Returns the cache."""
    from datetime import datetime, timezone

    entries = entries_only(items)
    no_doi = [e.get("Author", "?") for e in entries if not get_doi(e)]
    if no_doi:
        print(f"Note: {len(no_doi)} entry(ies) without a DOI will not be linked.")

    dois, seen = [], set()
    for e in entries:
        d = get_doi(e)
        if d and d not in seen:
            seen.add(d)
            dois.append(d)

    missing = [d for d in dois if d not in cache]
    if not missing:
        return cache
    if offline:
        print(f"Offline: {len(missing)} paper(s) not in cache, skipping fetch.")
        return cache

    print(f"Fetching references for {len(missing)} paper(s)…")
    for i, d in enumerate(missing):
        try:
            refs, source = fetch_references(d)
            cache[d] = {
                "references": refs,
                "source": source,
                "fetched_at": datetime.now(timezone.utc).isoformat(),
            }
            note = "" if refs else " (no open reference data available)"
            print(
                f"  [{i + 1}/{len(missing)}] {d}: {len(refs)} references"
                f" via {source}{note}"
            )
        except Exception as exc:
            print(f"  [{i + 1}/{len(missing)}] {d}: failed ({exc})")
        if i < len(missing) - 1:
            time.sleep(3.0)
    return cache


def build_citation_graph(items, cache) -> dict:
    """Build a JSON-serializable graph payload from entries and cached references."""
    entries = entries_only(items)
    nodes, doi_to_id, id_for = [], {}, {}
    for idx, e in enumerate(entries):
        nid = f"n{idx}"
        doi = get_doi(e)
        author = e.get("Author", "") or ""
        year = e.get("Date") or e.get("date") or e.get("DATE") or ""
        label = f"{author} ({year})" if year else author
        nodes.append(
            {
                "id": nid,
                "label": label or f"Paper {idx + 1}",
                "doi": doi,
                "journal": e.get("Journal") or e.get("journal") or "",
                "title": e.get("Title") or e.get("title") or "",
            }
        )
        id_for[idx] = nid
        if doi:
            doi_to_id[doi] = nid

    collection_edges, references = [], {}
    for idx, e in enumerate(entries):
        nid = id_for[idx]
        doi = get_doi(e)
        refs = cache.get(doi, {}).get("references", []) if doi else []
        ref_list = []
        for r in refs:
            rdoi = (r.get("doi") or "").lower()
            in_coll = bool(rdoi) and rdoi in doi_to_id
            ref_list.append(
                {
                    "doi": rdoi,
                    "title": r.get("title") or "",
                    "year": r.get("year"),
                    "author": r.get("author") or "",
                    "in_collection": in_coll,
                }
            )
            if in_coll and doi_to_id[rdoi] != nid:
                collection_edges.append([nid, doi_to_id[rdoi]])
        references[nid] = ref_list

    return {
        "nodes": nodes,
        "collection_edges": collection_edges,
        "references": references,
    }


GRAPH_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>__TITLE__ — citation graph</title>
<script src="https://unpkg.com/vis-network/standalone/umd/vis-network.min.js"></script>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    font-size: 16px; background: #f5f5f5; color: #222; padding: 1.5rem; }
  .header { display: flex; align-items: center; gap: 0.75rem; margin-bottom: 1rem; }
  h1 { font-size: 1.4rem; font-weight: 600; color: #333; }
  .controls { display: flex; gap: 0.75rem; align-items: center;
    margin-bottom: 1rem; flex-wrap: wrap; }
  select { padding: 7px 12px; border: 1px solid #ccc; border-radius: 6px;
    font-size: 14px; background: #fff; outline: none; max-width: 360px; }
  select:focus { border-color: #2d6a4f; }
  label { font-size: 0.85rem; color: #555; font-weight: 600; }
  #network { height: 74vh; background: #fff; border-radius: 8px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.12); }
  .count { margin-top: 0.75rem; font-size: 0.85rem; color: #888; }
  .legend { font-size: 0.8rem; color: #666; display: flex; gap: 1rem; align-items: center; }
  .dot { display: inline-block; width: 11px; height: 11px; border-radius: 50%;
    margin-right: 4px; vertical-align: middle; }
  .hint { font-size: 0.8rem; color: #999; }
  #detail { margin-top: 0.75rem; padding: 10px 14px; background: #fff;
    border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,0.12);
    font-size: 0.9rem; max-width: 720px; }
  #detail a { color: #2d6a4f; word-break: break-all; }
  #detail .doi-row { margin-top: 6px; }
  #detail button { margin-left: 8px; padding: 3px 10px; border: 1px solid #2d6a4f;
    background: #2d6a4f; color: #fff; border-radius: 5px; cursor: pointer;
    font-size: 0.8rem; }
  #detail .no-doi { color: #999; margin-top: 6px; }
</style>
</head>
<body>
  <div class="header"><h1>__TITLE__ — citation graph</h1></div>
  <div class="controls">
    <label for="view">View</label>
    <select id="view">
      <option value="collection">Collection network</option>
      <option value="paper">Single paper</option>
    </select>
    <span id="paper-wrap" style="display:none;">
      <label for="paper">Paper</label>
      <select id="paper"></select>
    </span>
    <span class="legend">
      <span><span class="dot" style="background:#2d6a4f;"></span>in collection</span>
      <span><span class="dot" style="background:#bbb;"></span>external</span>
    </span>
  </div>
  <div id="network"></div>
  <p class="count" id="count"></p>
  <p class="hint">Click any node to see its title and copy its DOI.</p>
  <div id="detail" style="display:none;"></div>

  <script>
    const GRAPH = __GRAPH_JSON__;
    const FOCUS = __FOCUS__;
    const COLL = "#2d6a4f", EXT = "#bbb", FOCUSC = "#d00000";
    const container = document.getElementById("network");
    const viewSel = document.getElementById("view");
    const paperSel = document.getElementById("paper");
    let network = null, currentNodes = null;

    function esc(s) {
      return String(s).replace(/[&<>"]/g, c =>
        ({ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;" }[c]));
    }

    function tip(title, doi) {
      return (title || "") + (doi ? "\\n" + doi : "");
    }

    GRAPH.nodes.forEach(n => {
      const o = document.createElement("option");
      o.value = n.id; o.textContent = n.label;
      paperSel.appendChild(o);
    });

    function draw() {
      let nodes = [], edges = [];
      if (viewSel.value === "collection") {
        nodes = GRAPH.nodes.map(n => ({
          id: n.id, label: n.label, color: COLL, shape: "dot", size: 14,
          doi: n.doi || "", refTitle: n.title || n.label,
          title: tip((n.title || n.label) + (n.journal ? "\\n" + n.journal : ""), n.doi)
        }));
        edges = GRAPH.collection_edges.map(([f, t]) => ({ from: f, to: t, arrows: "to" }));
        document.getElementById("count").textContent =
          nodes.length + " papers, " + edges.length + " citations among them";
      } else {
        const pid = paperSel.value;
        const focusNode = GRAPH.nodes.find(n => n.id === pid);
        const refs = GRAPH.references[pid] || [];
        nodes.push({ id: pid, label: focusNode.label, color: FOCUSC, shape: "dot",
          size: 22, doi: focusNode.doi || "", refTitle: focusNode.title || focusNode.label,
          title: tip(focusNode.title || focusNode.label, focusNode.doi) });
        let i = 0, inColl = 0;
        refs.forEach(r => {
          const rid = "r" + (i++);
          const lbl = (r.author || "?") + (r.year ? " (" + r.year + ")" : "");
          if (r.in_collection) inColl++;
          nodes.push({ id: rid, label: lbl,
            color: r.in_collection ? COLL : EXT, shape: "dot",
            size: r.in_collection ? 14 : 9,
            doi: r.doi || "", refTitle: r.title || lbl,
            title: tip(r.title || lbl, r.doi) });
          edges.push({ from: pid, to: rid, arrows: "to",
            color: { color: r.in_collection ? COLL : "#ddd" } });
        });
        document.getElementById("count").textContent =
          refs.length + " references (" + inColl + " in your collection)";
      }
      const data = { nodes: new vis.DataSet(nodes), edges: new vis.DataSet(edges) };
      currentNodes = data.nodes;
      document.getElementById("detail").style.display = "none";
      const options = {
        layout: { randomSeed: 42 },
        nodes: { font: { size: 14, color: "#222" }, borderWidth: 0 },
        edges: { color: { color: "#cbcbcb" }, smooth: { type: "continuous" }, width: 1 },
        physics: { stabilization: true,
          barnesHut: { gravitationalConstant: -8000, springLength: 120 } },
        interaction: { hover: true, tooltipDelay: 120 }
      };
      network = new vis.Network(container, data, options);
      network.on("click", showDetail);
    }

    function showDetail(params) {
      const panel = document.getElementById("detail");
      if (!params.nodes.length) { panel.style.display = "none"; return; }
      const n = currentNodes.get(params.nodes[0]);
      const doi = n.doi || "";
      let html = "<strong>" + esc(n.refTitle || n.label) + "</strong>";
      if (doi) {
        html += '<div class="doi-row">DOI: <a href="https://doi.org/'
          + encodeURIComponent(doi) + '" target="_blank" rel="noopener">'
          + esc(doi) + '</a><button id="copy-doi">Copy</button></div>';
      } else {
        html += '<div class="no-doi">No DOI available for this reference.</div>';
      }
      panel.innerHTML = html;
      panel.style.display = "block";
      const btn = document.getElementById("copy-doi");
      if (btn) btn.addEventListener("click", () => {
        navigator.clipboard.writeText(doi).then(
          () => { btn.textContent = "Copied!"; setTimeout(() => btn.textContent = "Copy", 1200); },
          () => { btn.textContent = "Copy failed"; }
        );
      });
    }

    viewSel.addEventListener("change", () => {
      document.getElementById("paper-wrap").style.display =
        viewSel.value === "paper" ? "" : "none";
      draw();
    });
    paperSel.addEventListener("change", draw);

    if (FOCUS) {
      viewSel.value = "paper";
      document.getElementById("paper-wrap").style.display = "";
      paperSel.value = FOCUS;
    }
    draw();
  </script>
</body>
</html>"""


def build_graph_html(items, filepath, focus=None, cache=None) -> str:
    """Render the interactive citation-graph page as a self-contained HTML string."""
    import json

    if cache is None:
        cache = load_graph_cache(graph_cache_path(filepath))
    graph = build_citation_graph(items, cache)
    title = Path(filepath).stem

    focus_id = ""
    if focus:
        fl = str(focus).strip().lower()
        for n in graph["nodes"]:
            if fl in n["label"].lower() or (n["title"] and fl in n["title"].lower()):
                focus_id = n["id"]
                break
        if not focus_id:
            print(f"Warning: no paper matched --graph-focus '{focus}'.")

    return (
        GRAPH_TEMPLATE.replace("__TITLE__", esc(title))
        .replace("__FOCUS__", json.dumps(focus_id))
        .replace("__GRAPH_JSON__", json.dumps(graph))
    )


# ── Zotero import ──────────────────────────────────────────────────────────────


ZOTERO_BASE = "http://localhost:23119/api/users/0"


def zotero_get(path: str):
    """GET a local Zotero API path and return parsed JSON (with response headers)."""
    import json
    from urllib.error import HTTPError, URLError
    from urllib.request import Request, urlopen

    url = f"{ZOTERO_BASE}{path}"
    req = Request(url, headers={"Accept": "application/json"})
    try:
        with urlopen(req, timeout=15) as resp:
            return json.load(resp), resp.headers
    except HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", "replace")
        except Exception:
            pass
        if e.code == 403 and "not enabled" in body.lower():
            raise SystemExit(
                "Zotero's local API is disabled. In Zotero open\n"
                "  Settings → Advanced → and enable\n"
                '  "Allow other applications on this computer to communicate '
                'with Zotero via the local API",\n'
                "then try again."
            )
        raise SystemExit(f"Zotero API error {e.code}: {body[:200]}")
    except URLError:
        raise SystemExit(
            "Could not reach Zotero at http://localhost:23119 — is Zotero running?"
        )


def _zotero_paged(path: str) -> list:
    """Fetch all results from a paginated Zotero collection/items endpoint."""
    sep = "&" if "?" in path else "?"
    results, start, limit = [], 0, 100
    while True:
        data, headers = zotero_get(f"{path}{sep}limit={limit}&start={start}")
        if not data:
            break
        results.extend(data)
        total = headers.get("Total-Results")
        if total is not None and len(results) >= int(total):
            break
        if len(data) < limit:
            break
        start += limit
    return results


def zotero_find_collection(name: str) -> str:
    """Return the collection key whose name matches `name` (case-insensitive)."""
    collections = _zotero_paged("/collections")
    target = name.strip().lower()
    matches = [
        c for c in collections if (c.get("data", {}).get("name", "")).lower() == target
    ]
    if not matches:
        names = sorted(c.get("data", {}).get("name", "") for c in collections)
        listing = "\n  ".join(names) if names else "(none found)"
        raise SystemExit(
            f"No Zotero collection named '{name}'. Available collections:\n  {listing}"
        )
    if len(matches) > 1:
        raise SystemExit(
            f"Multiple Zotero collections named '{name}'; rename one to disambiguate."
        )
    return matches[0]["key"]


def zotero_collection_items(key: str) -> list:
    """Return top-level items (no attachments/notes) of a collection."""
    return _zotero_paged(f"/collections/{key}/items/top")


def _zotero_author(creators: list) -> str:
    """Format a Zotero creators list into a short author label."""
    names = []
    for c in creators:
        if c.get("creatorType") not in (None, "author", "editor"):
            continue
        last = c.get("lastName") or c.get("name") or ""
        if last:
            names.append(last.strip())
    if not names:
        return ""
    if len(names) == 1:
        return names[0]
    if len(names) == 2:
        return f"{names[0]} & {names[1]}"
    return f"{names[0]} et al."


def zotero_item_to_entry(data: dict) -> Optional[dict]:
    """Map a Zotero item's `data` object to a reading-notes entry, or None to skip."""
    if data.get("itemType") in ("attachment", "note", "annotation"):
        return None
    year = ""
    m = re.search(r"\d{4}", str(data.get("date", "")))
    if m:
        year = m.group(0)
    return {
        "author": _zotero_author(data.get("creators", []) or []),
        "date": year,
        "journal": data.get("publicationTitle")
        or data.get("proceedingsTitle")
        or data.get("bookTitle")
        or "",
        "doi": get_doi({"doi": data.get("DOI", "")}),
        "title": data.get("title", "") or "",
        "zotero_key": data.get("key", "") or "",
    }


def existing_zotero_keys_and_dois(filepath: str) -> tuple[set, set]:
    """Collect zotero_key and normalized DOI values already in the YAML file."""
    if not Path(filepath).exists():
        return set(), set()
    keys, dois = set(), set()
    for e in entries_only(parse_yaml(filepath)):
        if e.get("zotero_key"):
            keys.add(e["zotero_key"])
        d = get_doi(e)
        if d:
            dois.add(d)
    return keys, dois


def _yaml_quote(value: str) -> str:
    """Double-quote a scalar so any YAML-special characters stay safe."""
    s = str(value).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{s}"'


def format_entry_yaml(entry: dict) -> str:
    """Render one entry as a 4/6-space-indented YAML block."""
    lines = [f"    - author: {_yaml_quote(entry.get('author', ''))}"]
    for field in ("date", "journal", "title"):
        val = entry.get(field, "")
        if val:
            lines.append(f"      {field}: {_yaml_quote(val)}")
    lines.append(f"      doi: {_yaml_quote(entry.get('doi', ''))}")
    lines.append(f"      zotero_key: {_yaml_quote(entry.get('zotero_key', ''))}")
    return "\n".join(lines)


def append_entries(filepath: str, collection_name: str, entries: list) -> None:
    """Append new entry blocks under a category, without rewriting existing bytes."""
    blocks = "\n\n".join(format_entry_yaml(e) for e in entries)
    cat_header = f"- category: {_yaml_quote(collection_name)}\n  entries:\n"

    path = Path(filepath)
    if not path.exists():
        path.write_text(cat_header + blocks + "\n", encoding="utf-8")
        return

    text = path.read_text(encoding="utf-8")
    lines = text.splitlines(keepends=True)

    # Find an existing category block matching this collection name.
    cat_idx = None
    pat = re.compile(r'^\s*-\s*category:\s*"?(.*?)"?\s*$')
    for i, line in enumerate(lines):
        m = pat.match(line)
        if m and m.group(1).strip() == collection_name:
            cat_idx = i
            break

    if cat_idx is None:
        sep = "" if text.endswith("\n") or not text else "\n"
        new_text = text + sep + "\n" + cat_header + blocks + "\n"
        path.write_text(new_text, encoding="utf-8")
        return

    # Insert before the next top-level list item (next category) or at EOF.
    insert_at = len(lines)
    for j in range(cat_idx + 1, len(lines)):
        if re.match(r"^-\s", lines[j]):
            insert_at = j
            break
    # Trim trailing blank lines within this block so spacing stays tidy.
    while insert_at > cat_idx + 1 and lines[insert_at - 1].strip() == "":
        insert_at -= 1
    snippet = "\n" + blocks + "\n"
    new_lines = lines[:insert_at] + [snippet] + lines[insert_at:]
    path.write_text("".join(new_lines), encoding="utf-8")


def import_from_zotero(filepath: str, collection_name: str) -> None:
    """Import a Zotero collection into the YAML file (append-only merge)."""
    key = zotero_find_collection(collection_name)
    raw = zotero_collection_items(key)
    candidates = []
    for item in raw:
        entry = zotero_item_to_entry(item.get("data", {}))
        if entry:
            candidates.append(entry)

    known_keys, known_dois = existing_zotero_keys_and_dois(filepath)
    new_entries, skipped = [], 0
    seen_keys = set()
    for e in candidates:
        k, d = e["zotero_key"], e["doi"]
        if (k and k in known_keys) or (d and d in known_dois) or (k and k in seen_keys):
            skipped += 1
            continue
        if k:
            seen_keys.add(k)
        new_entries.append(e)

    if not new_entries:
        print(f"Up to date: 0 new papers (skipped {skipped} already present).")
        return

    append_entries(filepath, collection_name, new_entries)
    no_doi = sum(1 for e in new_entries if not e["doi"])
    print(
        f"Added {len(new_entries)} new paper(s) to {filepath} "
        f"(skipped {skipped} already present)."
    )
    if no_doi:
        print(f"  {no_doi} of them have no DOI in Zotero — add one to link them.")


# ── Terminal rendering ─────────────────────────────────────────────────────────


def render_table(items, columns=None, sort_by=None, filter_by=None, max_width=30):
    result, sectioned = apply_filters(items, sort_by, filter_by)
    if result is None:
        return

    all_entries = entries_only(result) if sectioned else result
    cols = resolve_columns(all_entries, columns)

    def truncate(text):
        return text[: max_width - 1] + "…" if len(text) > max_width else text

    all_rows = [
        [truncate(strip_inline(e.get(col, ""))) for col in cols] for e in all_entries
    ]
    widths = [
        max(len(col), max((len(r[i]) for r in all_rows), default=0))
        for i, col in enumerate(cols)
    ]
    total_width = sum(widths) + 3 * len(widths) + 1

    def fmt_row(cells):
        return "| " + " | ".join(c.ljust(widths[i]) for i, c in enumerate(cells)) + " |"

    print(fmt_row(cols))
    print("|-" + "-+-".join("-" * w for w in widths) + "-|")

    if sectioned:
        for item in result:
            if item["_type"] == "section":
                label = f"  {item['title'].upper()}  "
                print("+" + label.center(total_width - 2, "-") + "+")
            elif item["_type"] == "subsection":
                label = f"  {item['title']}  "
                print("|" + label.center(total_width - 2, "·") + "|")
            else:
                print(
                    fmt_row([truncate(strip_inline(item.get(col, ""))) for col in cols])
                )
    else:
        for e in result:
            print(fmt_row([truncate(strip_inline(e.get(col, ""))) for col in cols]))


# ── XLSX export ────────────────────────────────────────────────────────────────


def write_xlsx(
    items, filepath, out_path=None, columns=None, sort_by=None, filter_by=None
):
    if not _OPENPYXL:
        raise SystemExit(
            "openpyxl is required for xlsx export: sudo pacman -S python-openpyxl"
        )

    result, sectioned = apply_filters(items, sort_by, filter_by)
    if result is None:
        return

    all_entries = entries_only(result) if sectioned else result
    cols = resolve_columns(all_entries, columns)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = Path(filepath).stem[:31]

    header_fill = PatternFill("solid", fgColor="2D6A4F")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="B7E4C7")
    section_font = Font(bold=True, color="1B4332")
    subsection_fill = PatternFill("solid", fgColor="D8F3DC")
    subsection_font = Font(bold=True, color="2D6A4F")

    # Header row
    for col_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False)

    row_idx = 2
    if sectioned:
        for item in result:
            if item["_type"] == "section":
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=1,
                    end_row=row_idx,
                    end_column=len(cols),
                )
                cell = ws.cell(row=row_idx, column=1, value=item["title"].upper())
                cell.fill = section_fill
                cell.font = section_font
            elif item["_type"] == "subsection":
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=1,
                    end_row=row_idx,
                    end_column=len(cols),
                )
                cell = ws.cell(row=row_idx, column=1, value=item["title"])
                cell.fill = subsection_fill
                cell.font = subsection_font
                cell.alignment = Alignment(indent=1)
            else:
                for col_idx, col in enumerate(cols, 1):
                    ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=strip_inline(item.get(col, "")),
                    )
            row_idx += 1
    else:
        for entry in result:
            for col_idx, col in enumerate(cols, 1):
                ws.cell(
                    row=row_idx, column=col_idx, value=strip_inline(entry.get(col, ""))
                )
            row_idx += 1

    # Auto-fit column widths (capped at 60)
    for col_idx, col in enumerate(cols, 1):
        max_len = len(col)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"

    dest = out_path or Path(filepath).with_suffix(".xlsx")
    wb.save(dest)
    print(f"Saved: {dest}")


# ── Live server ────────────────────────────────────────────────────────────────


def start_live_server(filepath, columns, sort_by, filter_by, port=8765):
    sse_clients = []
    lock = threading.Lock()

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, format, *args):
            pass

        def do_GET(self):
            if self.path in ("/", "/table"):
                items = (
                    parse_yaml(filepath)
                    if Path(filepath).suffix in (".yaml", ".yml")
                    else parse_org(filepath)
                )
                html = build_html(
                    items, filepath, columns, sort_by, filter_by, live=True
                )
                self._respond(200, "text/html", html.encode())

            elif self.path == "/sse":
                self.send_response(200)
                self.send_header("Content-Type", "text/event-stream")
                self.send_header("Cache-Control", "no-cache")
                self.send_header("Connection", "keep-alive")
                self.end_headers()
                with lock:
                    sse_clients.append(self.wfile)
                try:
                    while True:
                        time.sleep(1)
                except (BrokenPipeError, ConnectionResetError):
                    pass
                finally:
                    with lock:
                        if self.wfile in sse_clients:
                            sse_clients.remove(self.wfile)
            else:
                self._respond(404, "text/plain", b"Not found")

        def _respond(self, code, ctype, body):
            self.send_response(code)
            self.send_header("Content-Type", ctype)
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    def watch_file():
        last_mtime = os.path.getmtime(filepath)
        while True:
            time.sleep(0.5)
            try:
                mtime = os.path.getmtime(filepath)
                if mtime != last_mtime:
                    last_mtime = mtime
                    msg = b"data: reload\n\n"
                    with lock:
                        dead = []
                        for client in sse_clients:
                            try:
                                client.write(msg)
                                client.flush()
                            except Exception:
                                dead.append(client)
                        for d in dead:
                            sse_clients.remove(d)
            except FileNotFoundError:
                pass

    # ThreadingHTTPServer handles each connection in its own thread,
    # so the SSE connection no longer blocks /table requests.
    for attempt in range(10):
        try:
            server = ThreadingHTTPServer(("localhost", port + attempt), Handler)
            port = port + attempt
            break
        except OSError:
            if attempt == 9:
                raise
            continue
    threading.Thread(target=watch_file, daemon=True).start()

    url = f"http://localhost:{port}"
    print(f"Serving at {url}  (Ctrl+C to stop)")
    webbrowser.open(url)

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")


# ── CLI ────────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Render org properties as a table.")
    parser.add_argument("file", help="Path to the org file")
    parser.add_argument("--columns", "-c", nargs="+", help="Columns to display")
    parser.add_argument("--sort", "-s", help="Sort by property (e.g. Date)")
    parser.add_argument(
        "--filter", "-f", help="Filter by property=value (e.g. Journal=APSR)"
    )
    parser.add_argument(
        "--max-width",
        "-w",
        type=int,
        default=30,
        help="Max cell width in terminal (default: 30)",
    )
    parser.add_argument(
        "--html", action="store_true", help="Open as styled HTML table in browser"
    )
    parser.add_argument(
        "--live", action="store_true", help="Live update in browser as you edit"
    )
    parser.add_argument(
        "--port", type=int, default=8765, help="Port for live server (default: 8765)"
    )
    parser.add_argument(
        "--xlsx",
        metavar="OUT",
        nargs="?",
        const="",
        help="Export to xlsx (optional output path)",
    )
    parser.add_argument(
        "--graph",
        action="store_true",
        help="Open an interactive citation graph (requires DOIs)",
    )
    parser.add_argument(
        "--graph-focus",
        metavar="PAPER",
        help="Open the graph focused on one paper (matches Author/title)",
    )
    parser.add_argument(
        "--graph-offline",
        action="store_true",
        help="Build the graph from cache only, without calling the API",
    )
    parser.add_argument(
        "--from-zotero",
        metavar="COLLECTION",
        help="Append a Zotero collection's papers into the file (append-only)",
    )
    args = parser.parse_args()

    if args.from_zotero:
        import_from_zotero(args.file, args.from_zotero)
        return

    items = (
        parse_yaml(args.file)
        if Path(args.file).suffix in (".yaml", ".yml")
        else parse_org(args.file)
    )

    if args.graph or args.graph_focus or args.graph_offline:
        cache_path = graph_cache_path(args.file)
        cache = load_graph_cache(cache_path)
        fetch_all_references(items, cache, offline=args.graph_offline)
        save_graph_cache(cache_path, cache)
        html = build_graph_html(items, args.file, focus=args.graph_focus, cache=cache)
        tmp = tempfile.NamedTemporaryFile(
            delete=False, suffix=".html", mode="w", encoding="utf-8"
        )
        tmp.write(html)
        tmp.close()
        webbrowser.open(f"file://{tmp.name}")
        print(f"Opened: {tmp.name}")
    elif args.xlsx is not None:
        out = args.xlsx if args.xlsx else None
        write_xlsx(items, args.file, out, args.columns, args.sort, args.filter)
    elif args.live:
        start_live_server(args.file, args.columns, args.sort, args.filter, args.port)
    elif args.html:
        html = build_html(items, args.file, args.columns, args.sort, args.filter)
        tmp = tempfile.NamedTemporaryFile(
            delete=False, suffix=".html", mode="w", encoding="utf-8"
        )
        tmp.write(html)
        tmp.close()
        webbrowser.open(f"file://{tmp.name}")
        print(f"Opened: {tmp.name}")
    else:
        render_table(items, args.columns, args.sort, args.filter, args.max_width)


if __name__ == "__main__":
    main()
